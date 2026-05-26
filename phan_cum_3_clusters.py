"""
Phân cụm học lực sinh viên K58KTP sử dụng K-Means (k=3)
Bao gồm GPA, độ lệch chuẩn, số môn học, ...
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import io, warnings
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────────────────────
# BƯỚC 1: ĐỌC DỮ LIỆU
# ─────────────────────────────────────────────────────────────────────
FILE = r'C:\Users\nguye\OneDrive\Desktop\thuchanh\TỔNG HỢP ĐIỂM K58KTP.xlsx'

df_raw     = pd.read_excel(FILE, sheet_name=0, header=None)
mssv_list  = df_raw.iloc[1, 3:].tolist()
name_list  = df_raw.iloc[2, 3:].tolist()
subj_rows  = df_raw.iloc[4:, :]
n_students = len(mssv_list)

print(f'Số sinh viên: {n_students}')
print(f'Số môn học  : {len(subj_rows)}')

# ─────────────────────────────────────────────────────────────────────
# BƯỚC 2: TRÍCH XUẤT ĐẶC TRƯNG (CÓ THÊM ĐỘ LỆCH CHUẨN)
# ─────────────────────────────────────────────────────────────────────
student_data = []

for i, col_idx in enumerate(range(3, 3 + n_students)):
    scores = []
    for v in subj_rows.iloc[:, col_idx]:
        try:
            n = float(v)
            if 0 <= n <= 4:
                scores.append(n)
        except (TypeError, ValueError):
            pass
    
    if not scores:
        continue
    
    gpa = np.mean(scores)
    std_dev = np.std(scores)  # ← THÊM độ lệch chuẩn
    
    student_data.append({
        'STT'            : i + 1,
        'MSSV'           : mssv_list[i],
        'Họ và Tên'      : name_list[i],
        'Số môn'         : len(scores),
        'GPA'            : round(gpa, 4),
        'Độ lệch chuẩn'  : round(std_dev, 4),  # ← THÊM cột này
        'Môn xuất sắc'   : sum(1 for s in scores if s >= 3.6),
        'Môn giỏi'       : sum(1 for s in scores if 3.2 <= s < 3.6),
        'Môn yếu'        : sum(1 for s in scores if s < 2.0),
        'Tỉ lệ xuất sắc' : round(sum(1 for s in scores if s >= 3.6) / len(scores), 4),
    })

sv_df = pd.DataFrame(student_data)
print(f'Sinh viên hợp lệ: {len(sv_df)}')
print('\nDữ liệu mẫu:')
print(sv_df[['MSSV','Họ và Tên','Số môn','GPA','Độ lệch chuẩn','Môn xuất sắc']].head(10))

# ─────────────────────────────────────────────────────────────────────
# BƯỚC 3: K-MEANS CLUSTERING (k=3)
# ─────────────────────────────────────────────────────────────────────
# Chuẩn bị dữ liệu cho K-Means
X = sv_df[['GPA', 'Độ lệch chuẩn', 'Tỉ lệ xuất sắc']].values
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)

# Thực hiện K-Means với k=3
kmeans = KMeans(n_clusters=3, random_state=42, n_init=10)
clusters = kmeans.fit_predict(X_scaled)
sv_df['Cụm'] = clusters

# Gán tên nhóm dựa trên GPA trung bình của mỗi cụm
cluster_stats = sv_df.groupby('Cụm')['GPA'].agg(['mean', 'count']).sort_values('mean', ascending=False)
cluster_mapping = {idx: rank for rank, idx in enumerate(cluster_stats.index)}
cluster_names = {0: 'Nhóm cao', 1: 'Nhóm trung bình', 2: 'Nhóm thấp'}
sv_df['Học lực'] = sv_df['Cụm'].map({k: cluster_names[v] for k, v in cluster_mapping.items()})

print('\n📊 KẾT QUẢ K-MEANS (k=3):')
print(sv_df.groupby('Học lực')[['GPA', 'Độ lệch chuẩn']].agg(['min', 'max', 'mean', 'count']))

# Sắp xếp kết quả
order = ['Nhóm cao', 'Nhóm trung bình', 'Nhóm thấp']
sv_sorted = (
    sv_df
    .assign(_ord=sv_df['Học lực'].map({v: i for i, v in enumerate(order)}))
    .sort_values(['_ord', 'GPA'], ascending=[True, False])
    .reset_index(drop=True)
)
sv_sorted['STT'] = range(1, len(sv_sorted) + 1)

# ─────────────────────────────────────────────────────────────────────
# BƯỚC 4: BIỂU ĐỒ TRÒN
# ─────────────────────────────────────────────────────────────────────
COLORS = {
    'Nhóm cao'         : {'hex': '#1B5E20', 'border': '#A5D6A7', 'opx': 'FF1B5E20', 'font': 'FFFFFFFF'},
    'Nhóm trung bình'  : {'hex': '#1565C0', 'border': '#90CAF9', 'opx': 'FF1565C0', 'font': 'FFFFFFFF'},
    'Nhóm thấp'        : {'hex': '#E65100', 'border': '#FFCC80', 'opx': 'FFE65100', 'font': 'FFFFFFFF'},
}

counts     = [len(sv_df[sv_df['Học lực'] == g]) for g in order]
colors_hex = [COLORS[g]['hex'] for g in order]
total      = sum(counts)

fig, ax = plt.subplots(figsize=(7, 5), facecolor='white')
wedges, texts, autotexts = ax.pie(
    counts,
    colors=colors_hex,
    autopct=lambda p: f'{p:.1f}%\n({round(p*total/100):.0f} SV)',
    startangle=140,
    pctdistance=0.72,
    wedgeprops=dict(width=0.55, edgecolor='white', linewidth=2),
)
for at in autotexts:
    at.set_fontsize(9)
    at.set_color('white')
    at.set_fontweight('bold')

legend_labels = [
    f'{g}  —  {c} SV  (GPA tb {sv_df[sv_df["Học lực"]==g]["GPA"].mean():.2f})'
    for g, c in zip(order, counts)
]
patches = [mpatches.Patch(color=COLORS[g]['hex'], label=l)
           for g, l in zip(order, legend_labels)]
ax.legend(handles=patches, loc='lower center', bbox_to_anchor=(0.5, -0.18),
          ncol=2, fontsize=9, frameon=False)
ax.set_title('Phân cụm học lực sinh viên K58KTP\n(K-Means, k=3)',
             fontsize=12, fontweight='bold', color='#1A237E', pad=14)
plt.tight_layout()
# plt.show()

# Lưu vào buffer
img_buf = io.BytesIO()
fig.savefig(img_buf, format='png', dpi=150, bbox_inches='tight',
            facecolor='white', edgecolor='none')
img_buf.seek(0)
plt.close(fig)
print('✅ Biểu đồ đã lưu')

# ─────────────────────────────────────────────────────────────────────
# BƯỚC 5: XUẤT EXCEL
# ─────────────────────────────────────────────────────────────────────
thin  = Side(style='thin', color='FFBDBDBD')
BDR   = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR   = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
HFILL = PatternFill('solid', start_color='FF1A237E')
HFONT = Font(name='Arial', bold=True, color='FFFFFFFF', size=10)
DFONT = Font(name='Arial', size=9)

def hdr(ws, row, vals, widths):
    for j, (v, w) in enumerate(zip(vals, widths), 1):
        c = ws.cell(row=row, column=j, value=v)
        c.fill = HFILL; c.font = HFONT; c.alignment = CTR; c.border = BDR
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.row_dimensions[row].height = 30

def title_row(ws, text, n_cols, bg='FFE8EAF6', fg='FF1A237E', size=13):
    ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
    ws['A1'] = text
    ws['A1'].font      = Font(name='Arial', bold=True, size=size, color=fg)
    ws['A1'].fill      = PatternFill('solid', start_color=bg)
    ws['A1'].alignment = CTR
    ws.row_dimensions[1].height = 30

wb = Workbook()

# ── Sheet 1: Tổng hợp ──
ws1 = wb.active; ws1.title = 'Tổng hợp'
title_row(ws1, f'PHÂN CỤM HỌC LỰC K58KTP (K-Means k=3)  |  {len(sv_df)} SV', 11)
hdr(ws1, 2,
    ['STT','MSSV','Họ và Tên','Số môn','GPA','Độ lệch chuẩn',
     'Môn ≥3.6','Môn 3.2–3.6','Môn <2','Tỉ lệ XS (%)','Học lực'],
    [5,16,24,8,8,11,9,9,9,10,12])
for i, row in sv_sorted.iterrows():
    er = i + 3; hl = row['Học lực']
    gf     = PatternFill('solid', start_color=COLORS[hl]['opx'])
    gfont  = Font(name='Arial', bold=True, size=9, color=COLORS[hl]['font'])
    vals = [int(row['STT']), row['MSSV'], row['Họ và Tên'],
            int(row['Số môn']), round(row['GPA'],2), round(row['Độ lệch chuẩn'],2),
            int(row['Môn xuất sắc']), int(row['Môn giỏi']),
            int(row['Môn yếu']), f"{row['Tỉ lệ xuất sắc']*100:.1f}%", hl]
    for j, v in enumerate(vals, 1):
        c = ws1.cell(row=er, column=j, value=v)
        c.border = BDR; c.alignment = LEFT if j == 3 else CTR
        if j == 11: c.fill = gf; c.font = gfont
        else:
            c.font = DFONT
            if i % 2 == 0: c.fill = PatternFill('solid', start_color='FFF5F5F5')

# ── Sheet 2–4: Từng nhóm ──
for group in order:
    ws = wb.create_sheet(title=group)
    grp_df  = sv_sorted[sv_sorted['Học lực'] == group].reset_index(drop=True)
    gf      = PatternFill('solid', start_color=COLORS[group]['opx'])
    gf_font = Font(name='Arial', bold=True, size=9, color=COLORS[group]['font'])
    ws.merge_cells('A1:J1')
    ws['A1'] = (f"{group.upper()}  —  {len(grp_df)} SV  |  "
                f"GPA tb {grp_df['GPA'].mean():.2f}  "
                f"Độ lệch chuẩn tb {grp_df['Độ lệch chuẩn'].mean():.2f}")
    ws['A1'].font = gf_font; ws['A1'].fill = gf
    ws['A1'].alignment = CTR; ws.row_dimensions[1].height = 28
    hdr(ws, 2,
        ['STT','MSSV','Họ và Tên','Số môn','GPA','Độ lệch chuẩn',
         'Môn ≥3.6','Môn 3.2–3.6','Môn <2','Tỉ lệ XS (%)'],
        [5,16,26,8,8,11,10,10,8,11])
    for j in range(1, 11):
        ws.cell(2, j).fill = gf; ws.cell(2, j).font = gf_font
    for i, row in grp_df.iterrows():
        er = i + 3
        vals = [i+1, row['MSSV'], row['Họ và Tên'],
                int(row['Số môn']), round(row['GPA'],2), round(row['Độ lệch chuẩn'],2),
                int(row['Môn xuất sắc']), int(row['Môn giỏi']),
                int(row['Môn yếu']), f"{row['Tỉ lệ xuất sắc']*100:.1f}%"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=er, column=j, value=v)
            c.font = DFONT; c.border = BDR
            c.alignment = LEFT if j == 3 else CTR
            if i % 2 == 0: c.fill = PatternFill('solid', start_color='FFF5F5F5')

# ── Sheet 5: Thống kê ──
ws_s = wb.create_sheet('Thống kê')
title_row(ws_s, 'THỐNG KÊ PHÂN CỤM HỌC LỰC (K-Means k=3)', 7)
hdr(ws_s, 2,
    ['Học lực','Số SV','% SV','GPA TB','Độ lệch chuẩn','GPA cao nhất','GPA thấp nhất'],
    [14,8,10,12,14,14,14])
for i, g in enumerate(order):
    grp = sv_df[sv_df['Học lực'] == g]; er = i + 3
    gf2      = PatternFill('solid', start_color=COLORS[g]['opx'])
    gf2_font = Font(name='Arial', bold=True, size=9, color=COLORS[g]['font'])
    vals = [g, len(grp), f"{len(grp)/len(sv_df)*100:.1f}%",
            round(grp['GPA'].mean(),2), round(grp['Độ lệch chuẩn'].mean(),2),
            round(grp['GPA'].max(),2), round(grp['GPA'].min(),2)]
    for j, v in enumerate(vals, 1):
        c = ws_s.cell(row=er, column=j, value=v)
        c.border = BDR; c.alignment = CTR
        c.font = gf2_font if j == 1 else DFONT
        if j == 1: c.fill = gf2

img_xl        = XLImage(img_buf)
img_xl.width  = 480
img_xl.height = 340
ws_s.add_image(img_xl, 'A6')

OUT = 'PHAN_CUM_HOC_LUC_K58KTP_k3.xlsx'
wb.save(OUT)
print(f'\n✅ Đã lưu: {OUT}')
print('\n📈 THỐNG KÊ CUỐI CÙNG:')
for g in order:
    grp = sv_df[sv_df['Học lực'] == g]
    print(f'  {g:15s}: {len(grp):3d} SV  |  GPA {grp["GPA"].mean():.2f} ({grp["GPA"].min():.2f}–{grp["GPA"].max():.2f})  |  Độ lệch chuẩn {grp["Độ lệch chuẩn"].mean():.2f}')
